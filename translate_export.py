import json
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from deep_translator import GoogleTranslator

INPUT_FILE = "edit vs original 0507.xlsx"
OUTPUT_FILE = "SA_edits_translated.xlsx"

translator = GoogleTranslator(source='zh-CN', target='en')

def translate(text):
    if not text or not isinstance(text, str):
        return text
    # Skip if mostly non-Chinese (rough heuristic: <10% CJK chars)
    cjk = sum(1 for c in text if '一' <= c <= '鿿')
    if cjk < 2:
        return text
    try:
        # Google Translate has a 5000 char limit per call
        if len(text) <= 4900:
            return translator.translate(text)
        # Chunk if too long
        chunks = [text[i:i+4900] for i in range(0, len(text), 4900)]
        return ' '.join(translator.translate(c) for c in chunks)
    except Exception as e:
        print(f"  [warn] Translation failed: {e}")
        time.sleep(2)
        try:
            return translator.translate(text[:4900])
        except:
            return text  # fallback to original

wb_in = openpyxl.load_workbook(INPUT_FILE)
ws_in = wb_in['Sheet1']

headers = [ws_in.cell(1, col).value for col in range(1, ws_in.max_column + 1)]
col_idx = {h: i+1 for i, h in enumerate(headers)}

# Collect is_change=true rows
rows_to_export = []
for row in range(2, ws_in.max_row + 1):
    val = ws_in.cell(row, col_idx['copy_to_chat_history']).value
    if not val:
        continue
    try:
        data = json.loads(val)
        latest = data.get('latest_version', 'v1')
        lv = data.get(latest, {})
        if lv.get('is_change') is True:
            rows_to_export.append((row, data, latest, lv))
    except:
        pass

print(f"Found {len(rows_to_export)} is_change=True rows")

# Build output columns
out_headers = [
    'ai_action_plan_id',
    'seller_id',
    'unify_id',
    'use_case',
    'country',
    'created',
    'send_time',
    'conversation_starter_subject_original',
    'conversation_starter_subject_translated',
    'description_original',
    'description_translated',
    'ai_message_original',
    'ai_message_translated',
    'sent_message_original',
    'sent_message_translated',
    'num_versions',
]

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "SA Edits (Translated)"

# Style header row
header_fill = PatternFill("solid", fgColor="1F3864")
header_font = Font(bold=True, color="FFFFFF", size=11)
for col, h in enumerate(out_headers, 1):
    cell = ws_out.cell(1, col, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(wrap_text=True, vertical='center')

# Alternate row fills
fill_a = PatternFill("solid", fgColor="EEF2FF")
fill_b = PatternFill("solid", fgColor="FFFFFF")

total = len(rows_to_export)
for i, (row, data, latest, lv) in enumerate(rows_to_export):
    print(f"  [{i+1}/{total}] Translating row {row}...")

    def get(col_name):
        return ws_in.cell(row, col_idx[col_name]).value

    subject_orig   = get('conversation_starter_subject') or ''
    desc_orig      = get('description') or ''
    ai_msg_orig    = get('conversation_starter_message') or ''
    sent_msg_orig  = lv.get('send_content', '')
    send_time      = lv.get('send_time', '')
    num_versions   = len([k for k in data if k != 'latest_version'])

    subject_en  = translate(subject_orig)
    desc_en     = translate(desc_orig)
    ai_msg_en   = translate(ai_msg_orig)
    sent_msg_en = translate(sent_msg_orig)

    out_row = [
        get('ai_action_plan_id'),
        get('seller_id'),
        get('unify_id'),
        get('use_case'),
        get('country'),
        get('created'),
        send_time,
        subject_orig,
        subject_en,
        desc_orig,
        desc_en,
        ai_msg_orig,
        ai_msg_en,
        sent_msg_orig,
        sent_msg_en,
        num_versions,
    ]

    r = i + 2
    fill = fill_a if i % 2 == 0 else fill_b
    for col, val in enumerate(out_row, 1):
        cell = ws_out.cell(r, col, val)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Small pause to avoid rate limiting
    time.sleep(0.3)

# Auto-width columns (capped)
for col in ws_out.columns:
    max_len = max((len(str(c.value)) if c.value else 0) for c in col)
    ws_out.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

wb_out.save(OUTPUT_FILE)
print(f"\nDone! Saved to {OUTPUT_FILE}")
