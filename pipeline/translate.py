"""
Filter the raw xlsx to is_change=True rows, translate Chinese → English.
Accepts a BytesIO (from Streamlit uploader) instead of a hardcoded filename.
Returns (df_edited, df_full) DataFrames.
"""
import json
import time
import io
import pandas as pd
import openpyxl
from deep_translator import GoogleTranslator

translator = GoogleTranslator(source='zh-CN', target='en')


def _translate(text):
    if not text or not isinstance(text, str):
        return text
    cjk = sum(1 for c in text if '一' <= c <= '鿿')
    if cjk < 2:
        return text
    try:
        if len(text) <= 4900:
            return translator.translate(text)
        chunks = [text[i:i+4900] for i in range(0, len(text), 4900)]
        return ' '.join(translator.translate(c) for c in chunks)
    except Exception:
        time.sleep(2)
        try:
            return translator.translate(text[:4900])
        except Exception:
            return text


def run(raw_bytes: io.BytesIO, progress_callback=None):
    """
    Parameters
    ----------
    raw_bytes       : BytesIO of the uploaded xlsx file
    progress_callback : callable(current, total) for progress bar updates

    Returns
    -------
    df_edited : DataFrame — 181-ish rows that were edited, with translated columns
    df_full   : DataFrame — all rows from the raw file (for KPI denominators)
    """
    wb = openpyxl.load_workbook(raw_bytes)
    # Try Sheet1 first, fall back to active sheet
    ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    # Build df_full for KPI denominators
    full_rows = []
    edited_source = []
    for row in range(2, ws.max_row + 1):
        def get(col_name, r=row):
            return ws.cell(r, col_idx[col_name]).value if col_name in col_idx else None

        uc  = get('use_case')
        sid = get('seller_id')
        cth = get('copy_to_chat_history')
        full_rows.append({'use_case': uc, 'seller_id': sid, 'copy_to_chat_history': cth})

        if not cth:
            continue
        try:
            data   = json.loads(cth)
            latest = data.get('latest_version', 'v1')
            lv     = data.get(latest, {})
            if lv.get('is_change') is True:
                edited_source.append((row, data, latest, lv))
        except Exception:
            pass

    df_full = pd.DataFrame(full_rows)

    out_headers = [
        'ai_action_plan_id', 'seller_id', 'unify_id', 'use_case', 'country',
        'created', 'send_time',
        'conversation_starter_subject_original', 'conversation_starter_subject_translated',
        'description_original', 'description_translated',
        'ai_message_original', 'ai_message_translated',
        'sent_message_original', 'sent_message_translated',
        'num_versions',
    ]

    total = len(edited_source)
    out_rows = []
    for i, (row, data, latest, lv) in enumerate(edited_source):
        if progress_callback:
            progress_callback(i, total)

        def get(col_name, r=row):
            return ws.cell(r, col_idx[col_name]).value if col_name in col_idx else None

        subject_orig  = get('conversation_starter_subject') or ''
        desc_orig     = get('description') or ''
        ai_msg_orig   = get('conversation_starter_message') or ''
        sent_msg_orig = lv.get('send_content', '')
        send_time     = lv.get('send_time', '')
        num_versions  = len([k for k in data if k != 'latest_version'])

        subject_en  = _translate(subject_orig)
        desc_en     = _translate(desc_orig)
        ai_msg_en   = _translate(ai_msg_orig)
        sent_msg_en = _translate(sent_msg_orig)

        out_rows.append({
            'ai_action_plan_id':                          get('ai_action_plan_id'),
            'seller_id':                                  get('seller_id'),
            'unify_id':                                   get('unify_id'),
            'use_case':                                   get('use_case'),
            'country':                                    get('country'),
            'created':                                    get('created'),
            'send_time':                                  send_time,
            'conversation_starter_subject_original':      subject_orig,
            'conversation_starter_subject_translated':    subject_en,
            'description_original':                       desc_orig,
            'description_translated':                     desc_en,
            'ai_message_original':                        ai_msg_orig,
            'ai_message_translated':                      ai_msg_en,
            'sent_message_original':                      sent_msg_orig,
            'sent_message_translated':                    sent_msg_en,
            'num_versions':                               num_versions,
        })

        time.sleep(0.3)

    if progress_callback:
        progress_callback(total, total)

    df_edited = pd.DataFrame(out_rows, columns=out_headers)
    return df_edited, df_full
