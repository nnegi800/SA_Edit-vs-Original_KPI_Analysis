"""
Rebuild SA_edits_translated.xlsx using xlsxwriter (better Excel compatibility).
Rich text: red = removed from AI msg, bold = added in sent msg.
"""
import difflib, json, openpyxl, xlsxwriter
from openpyxl.cell.rich_text import CellRichText

SRC_FILE = "SA_edits_translated.xlsx"   # openpyxl-written source (has the data)
OUT_FILE = "SA_edits_final.xlsx"

# ── read all data from the broken file ───────────────────────────────────────
wb = openpyxl.load_workbook(SRC_FILE, rich_text=True)
ws = wb.active

def plain(v):
    if isinstance(v, CellRichText):
        return "".join(str(b) for b in v)
    return str(v).strip() if v else ""

headers = [plain(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
col = {h: i+1 for i, h in enumerate(headers)}

rows = []
for r in range(2, ws.max_row + 1):
    rows.append({h: plain(ws.cell(r, col[h]).value) for h in headers})

wb.close()
print(f"Read {len(rows)} rows, {len(headers)} columns")

# ── diff helpers ──────────────────────────────────────────────────────────────
def char_diff(a, b):
    sm = difflib.SequenceMatcher(None, a, b, autojunk=False)
    segs_a, segs_b = [], []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == 'equal':
            segs_a.append(('eq', a[i1:i2]))
            segs_b.append(('eq', b[j1:j2]))
        elif tag == 'delete':
            segs_a.append(('del', a[i1:i2]))
        elif tag == 'insert':
            segs_b.append(('ins', b[j1:j2]))
        elif tag == 'replace':
            segs_a.append(('del', a[i1:i2]))
            segs_b.append(('ins', b[j1:j2]))
    return segs_a, segs_b

# ── build xlsxwriter workbook ─────────────────────────────────────────────────
wb_out = xlsxwriter.Workbook(OUT_FILE)
ws_out = wb_out.add_worksheet("SA Edits")

# formats
hdr_fmt  = wb_out.add_format({'bold': True, 'font_color': '#FFFFFF', 'bg_color': '#1F3864',
                               'text_wrap': True, 'valign': 'vcenter', 'font_size': 11})
fill_a   = '#EEF2FF'
fill_b   = '#FFFFFF'
def row_fmt(even, extra=None):
    props = {'text_wrap': True, 'valign': 'top', 'font_size': 11,
             'bg_color': fill_a if even else fill_b}
    if extra:
        props.update(extra)
    return wb_out.add_format(props)

# pre-make base formats (normal / red / bold) for each row parity
fmt_cache = {}
def get_fmt(even, style='normal'):
    key = (even, style)
    if key not in fmt_cache:
        bg = fill_a if even else fill_b
        props = {'text_wrap': True, 'valign': 'top', 'font_size': 11, 'bg_color': bg}
        if style == 'red':
            props['font_color'] = '#CC0000'
        elif style == 'bold':
            props['bold'] = True
        fmt_cache[key] = wb_out.add_format(props)
    return fmt_cache[key]

def build_rich(segs, mode, even):
    """
    mode='ai'   → normal + red(del)
    mode='sent' → normal + bold(ins)
    Returns list ready for ws_out.write_rich_string(), or plain string if no diff.
    """
    parts = []
    has_markup = False
    for tag, text in segs:
        if not text:
            continue
        if mode == 'ai' and tag == 'del':
            parts += [get_fmt(even, 'red'), text]
            has_markup = True
        elif mode == 'sent' and tag == 'ins':
            parts += [get_fmt(even, 'bold'), text]
            has_markup = True
        else:
            parts += [get_fmt(even, 'normal'), text]
    return parts if has_markup else None

# ── write header ──────────────────────────────────────────────────────────────
for ci, h in enumerate(headers):
    ws_out.write(0, ci, h, hdr_fmt)

# column widths
widths = {
    'seller_id': 12, 'unify_id': 18, 'use_case': 22, 'created': 22, 'send_time': 20,
    'time_to_send (days)': 14,
    'conversation_starter_subject_original': 30, 'conversation_starter_subject_translated': 30,
    'description_original': 40, 'description_translated': 40,
    'ai_message_original': 50, 'ai_message_translated': 50,
    'sent_message_original': 50, 'sent_message_translated': 50,
    'change_summary': 60,
}
for ci, h in enumerate(headers):
    ws_out.set_column(ci, ci, widths.get(h, 18))
ws_out.set_row(0, 30)

# columns that get rich text diff
AI_ORIG_H   = 'ai_message_original'
AI_EN_H     = 'ai_message_translated'
SENT_ORIG_H = 'sent_message_original'
SENT_EN_H   = 'sent_message_translated'

DIFF_COLS = {AI_ORIG_H, AI_EN_H, SENT_ORIG_H, SENT_EN_H}

# ── write data rows ───────────────────────────────────────────────────────────
for ri, row in enumerate(rows):
    even = ri % 2 == 0
    xlsx_row = ri + 1   # 0-indexed, row 0 = header
    ws_out.set_row(xlsx_row, None, None, {'level': 0})

    # compute diffs once per row
    segs_ai_orig,   segs_sent_orig = char_diff(row[AI_ORIG_H],   row[SENT_ORIG_H])
    segs_ai_en,     segs_sent_en   = char_diff(row[AI_EN_H],     row[SENT_EN_H])

    rich_map = {
        AI_ORIG_H:   build_rich(segs_ai_orig,   'ai',   even),
        AI_EN_H:     build_rich(segs_ai_en,      'ai',   even),
        SENT_ORIG_H: build_rich(segs_sent_orig,  'sent', even),
        SENT_EN_H:   build_rich(segs_sent_en,    'sent', even),
    }

    for ci, h in enumerate(headers):
        val = row[h]
        base_fmt = get_fmt(even, 'normal')

        if h in DIFF_COLS and rich_map[h]:
            # write_rich_string needs at least one format+string pair + a cell_format at end
            ws_out.write_rich_string(xlsx_row, ci, *rich_map[h], base_fmt)
        elif h == 'time_to_send (days)':
            try:
                ws_out.write_number(xlsx_row, ci, float(val), base_fmt)
            except:
                ws_out.write(xlsx_row, ci, val, base_fmt)
        else:
            ws_out.write(xlsx_row, ci, val, base_fmt)

    if (ri + 1) % 30 == 0 or ri == 0:
        print(f"  [{ri+1}/{len(rows)}] written")

wb_out.close()
print(f"\nDone — {OUT_FILE}")
