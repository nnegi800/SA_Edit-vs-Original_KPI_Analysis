"""
SA Edit Analysis — Streamlit UI
Upload the raw xlsx → pipeline runs → KPI dashboard with evidence toggles.
"""
import io
import sys
import os

import pandas as pd
import streamlit as st

# Allow imports from this directory
sys.path.insert(0, os.path.dirname(__file__))

from pipeline import translate, enrich, kpis, evidence, export

st.set_page_config(
    page_title="SA Edit Analysis",
    page_icon="📊",
    layout="wide",
)

# ── minimal custom styles ──────────────────────────────────────────────────────
st.markdown("""
<style>
.banner-metric { text-align: center; padding: 12px 0; }
.banner-metric .value { font-size: 2.2rem; font-weight: 700; color: #1F3864; }
.banner-metric .label { font-size: 0.85rem; color: #666; margin-top: 2px; }
.section-header {
    background: #1F3864; color: white; font-weight: 700;
    padding: 6px 12px; border-radius: 4px; margin: 18px 0 6px 0;
    font-size: 0.9rem; letter-spacing: 0.04em;
}
.kpi-row { padding: 4px 0; }
div[data-testid="stExpander"] > div:first-child {
    font-size: 0.82rem; color: #555;
}
</style>
""", unsafe_allow_html=True)


# ── session state defaults ─────────────────────────────────────────────────────
for key in ('stage', 'df_edited', 'df_full', 'kpi_results', 'coloured_bytes', 'kpi_bytes'):
    if key not in st.session_state:
        st.session_state[key] = None
if st.session_state.stage is None:
    st.session_state.stage = 'upload'


# ══════════════════════════════════════════════════════════════════════════════
# STAGE 1 — UPLOAD
# ══════════════════════════════════════════════════════════════════════════════
def show_upload():
    st.title("SA Edit Pattern Analysis")
    st.markdown("Upload the raw data file to run the full analysis pipeline.")

    uploaded = st.file_uploader(
        "Drop your raw xlsx file here",
        type=["xlsx"],
        help="The file should be the full export of AI-generated messages (e.g. 'edit vs original 0507.xlsx')",
    )

    if uploaded:
        raw_bytes = io.BytesIO(uploaded.read())
        raw_bytes.seek(0)

        import openpyxl
        wb_prev = openpyxl.load_workbook(raw_bytes, read_only=True)
        ws_prev = wb_prev['Sheet1'] if 'Sheet1' in wb_prev.sheetnames else wb_prev.active
        row_count = ws_prev.max_row - 1
        col_count = ws_prev.max_column
        wb_prev.close()
        raw_bytes.seek(0)

        st.caption(f"{row_count:,} rows · {col_count} columns")

        if st.button("Run Analysis", type="primary", use_container_width=True):
            st.session_state.raw_bytes = raw_bytes
            st.session_state.stage = 'processing'
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# STAGE 2 — PROCESSING
# ══════════════════════════════════════════════════════════════════════════════
def show_processing():
    st.title("Running Analysis Pipeline")

    step1 = st.empty()
    prog1 = st.progress(0)
    step2 = st.empty()
    prog2 = st.empty()
    step3 = st.empty()

    # ── Step 1: filter + translate ──
    step1.markdown("**Step 1 / 3** — Filtering edited messages & translating Chinese → English...")

    def translate_progress(current, total):
        if total > 0:
            prog1.progress(current / total, text=f"{current} / {total} rows translated")

    df_edited, df_full = translate.run(
        st.session_state.raw_bytes,
        progress_callback=translate_progress,
    )
    prog1.progress(1.0, text=f"{len(df_edited)} edited rows found and translated")
    step1.markdown(f"**Step 1 / 3** — Done. Found **{len(df_edited)}** edited messages out of **{len(df_full)}** total.")

    # ── Step 2: enrich diffs ──
    step2.markdown("**Step 2 / 3** — Computing character-level diffs and change summaries...")
    prog2_bar = st.progress(0)

    total_rows = len(df_edited)
    processed  = [0]

    def enrich_progress(current, total):
        if total > 0:
            prog2_bar.progress(current / total, text=f"{current} / {total} rows enriched")
        processed[0] = current

    df_enriched = enrich.run(df_edited, progress_callback=enrich_progress)
    prog2_bar.progress(1.0, text=f"{total_rows} rows enriched")
    step2.markdown(f"**Step 2 / 3** — Done. Diffs and change summaries added.")

    # ── Step 3: KPIs ──
    step3.markdown("**Step 3 / 3** — Computing KPIs...")
    kpi_results = kpis.run(df_enriched, df_full)
    step3.markdown("**Step 3 / 3** — Done. KPI dashboard ready.")

    # ── pre-generate download bytes ──
    coloured_bytes = export.to_coloured_xlsx(df_enriched)
    kpi_bytes      = export.to_kpi_xlsx(kpi_results)

    st.session_state.df_edited    = df_enriched
    st.session_state.df_full      = df_full
    st.session_state.kpi_results  = kpi_results
    st.session_state.coloured_bytes = coloured_bytes
    st.session_state.kpi_bytes    = kpi_bytes
    st.session_state.stage        = 'dashboard'

    st.success("Analysis complete!")
    st.button("View KPI Dashboard →", type="primary", on_click=lambda: st.rerun())


# ══════════════════════════════════════════════════════════════════════════════
# STAGE 3 — KPI DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
def show_dashboard():
    kpi_results = st.session_state.kpi_results
    df_edited   = st.session_state.df_edited

    overall = kpi_results['overall']

    # ── banner ──
    st.title("SA Edit Pattern Analysis")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f"""<div class="banner-metric">
            <div class="value">{overall.get('total_generated', '—'):,}</div>
            <div class="label">AI Messages Generated</div></div>""", unsafe_allow_html=True)
    with c2:
        st.markdown(f"""<div class="banner-metric">
            <div class="value">{overall.get('total_edited', '—'):,}</div>
            <div class="label">Messages Edited by SAs</div></div>""", unsafe_allow_html=True)
    with c3:
        rate_raw = overall.get('edit_rate', '—')
        rate_display = rate_raw.split('%')[0] + '%' if '%' in str(rate_raw) else rate_raw
        st.markdown(f"""<div class="banner-metric">
            <div class="value">{rate_display}</div>
            <div class="label">Overall Edit Rate</div></div>""", unsafe_allow_html=True)
    with c4:
        st.markdown(f"""<div class="banner-metric">
            <div class="value">{overall.get('sa_count', '—')}</div>
            <div class="label">Sales Associates</div></div>""", unsafe_allow_html=True)

    st.divider()

    # ── use-case tabs ──
    tab_labels = ["Overall", "Emotional Bonding", "Task", "Product Storytelling"]
    tabs = st.tabs(tab_labels)

    for tab, tab_label in zip(tabs, tab_labels):
        with tab:
            uc_key    = None if tab_label == "Overall" else tab_label
            agg       = overall if tab_label == "Overall" else kpi_results.get(tab_label, {})
            _render_kpi_sections(kpi_results['sections'], agg, df_edited, uc_key, kpi_results['sa_table'])

    st.divider()

    # ── download buttons ──
    st.subheader("Download Files")
    dl1, dl2 = st.columns(2)
    with dl1:
        st.download_button(
            label="Download SA_edits_coloured.xlsx",
            data=st.session_state.coloured_bytes.getvalue(),
            file_name="SA_edits_coloured.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.caption("181 edited rows with red/bold diff highlights and change summaries")
    with dl2:
        st.download_button(
            label="Download KPI_Summary.xlsx",
            data=st.session_state.kpi_bytes.getvalue(),
            file_name="KPI_Summary.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.caption("All KPI metrics broken out by use case")

    # ── restart ──
    st.divider()
    if st.button("Upload a new file", use_container_width=False):
        for key in ('stage', 'df_edited', 'df_full', 'kpi_results', 'coloured_bytes', 'kpi_bytes', 'raw_bytes'):
            st.session_state[key] = None
        st.session_state.stage = 'upload'
        st.rerun()


def _render_kpi_sections(sections, agg, df_edited, use_case_label, sa_table):
    """Render all KPI sections with evidence expanders for the active tab."""
    for section_label, kpi_items in sections:
        st.markdown(f'<div class="section-header">{section_label}</div>', unsafe_allow_html=True)

        is_sa_section = "SA" in section_label

        for label, key in kpi_items:
            value = agg.get(key, '—')
            if value is None:
                value = '—'

            col_val, col_exp = st.columns([3, 1])
            with col_val:
                st.markdown(f"**{label}**")
                st.write(value)

            with col_exp:
                # Determine evidence type
                if key in evidence.SA_DIST_KEYS:
                    # SA-level evidence
                    with st.expander(f"Show evidence ({len(sa_table)} SAs)"):
                        if use_case_label:
                            UC_MAP = {
                                'Emotional Bonding':    'ai360_action_plan_emotional_bonding',
                                'Task':                 'ai360_action_plan_task',
                                'Product Storytelling': 'ai360_action_plan_product_storytelling',
                            }
                            raw_uc  = UC_MAP.get(use_case_label)
                            edited_sids = set(df_edited[df_edited['use_case'] == raw_uc]['seller_id'].unique()) if raw_uc else set()
                            sa_filtered = sa_table[sa_table['seller_id'].isin(edited_sids)] if edited_sids else sa_table
                            st.dataframe(sa_filtered, height=280, use_container_width=True)
                        else:
                            st.dataframe(sa_table, height=280, use_container_width=True)

                elif key in evidence.NO_EVIDENCE_KEYS:
                    st.caption("—")

                else:
                    evidence_rows = evidence.get_rows(key, df_edited, use_case_label)
                    if evidence_rows is not None and len(evidence_rows) > 0:
                        with st.expander(f"Show evidence ({len(evidence_rows)} rows)"):
                            st.dataframe(
                                evidence_rows,
                                height=min(300, 50 + len(evidence_rows) * 35),
                                use_container_width=True,
                            )
                    elif evidence_rows is not None and len(evidence_rows) == 0:
                        st.caption("No matching rows")
                    else:
                        st.caption("—")

            st.markdown("---")


# ══════════════════════════════════════════════════════════════════════════════
# ROUTER
# ══════════════════════════════════════════════════════════════════════════════
stage = st.session_state.stage

if stage == 'upload':
    show_upload()
elif stage == 'processing':
    show_processing()
elif stage == 'dashboard':
    show_dashboard()
