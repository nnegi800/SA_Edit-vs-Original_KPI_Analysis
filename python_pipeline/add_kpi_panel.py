"""
Append a KPI summary panel to the right of SA_edits_final.xlsx.
Rows = KPI metrics.  Columns = Overall + one per use_case.
"""
import difflib, re, openpyxl, xlsxwriter
from openpyxl.cell.rich_text import CellRichText
from collections import defaultdict

SRC      = "SA_edits_translated.xlsx"      # 181 edited rows with change_summary
SRC_ORIG = "edit vs original 0507.xlsx"   # full dataset for edit-rate denominator
OUT      = "KPI_Summary.xlsx"

# ── helpers ───────────────────────────────────────────────────────────────────
def plain(v):
    if isinstance(v, CellRichText):
        return "".join(str(b) for b in v)
    return str(v).strip() if v else ""

EMOJI_RE = re.compile(
    "[\U0001F600-\U0001F64F"
    "\U0001F300-\U0001F5FF"
    "\U0001F680-\U0001F6FF"
    "\U0001F1E0-\U0001F1FF"
    "\U00002600-\U000027BF"
    "\U0001F900-\U0001F9FF"
    "\U0001FA00-\U0001FA6F"
    "\U0001FA70-\U0001FAFF"
    "\U00002702-\U000027B0]+",
    flags=re.UNICODE
)

def has_emoji(text):
    return bool(EMOJI_RE.search(text))

def added_segments(a, b):
    """Return only the inserted characters (b minus a)."""
    sm = difflib.SequenceMatcher(None, a, b, autojunk=False)
    added = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag in ('insert', 'replace'):
            added.append(b[j1:j2])
    return "".join(added)

def removed_segments(a, b):
    sm = difflib.SequenceMatcher(None, a, b, autojunk=False)
    removed = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag in ('delete', 'replace'):
            removed.append(a[i1:i2])
    return "".join(removed)

# ── read data ─────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(SRC, rich_text=True)
ws = wb.active
headers = [plain(ws.cell(1,c).value) for c in range(1, ws.max_column+1)]
col = {h: i+1 for i,h in enumerate(headers)}

records = []
for r in range(2, ws.max_row+1):
    records.append({h: plain(ws.cell(r, col[h]).value) for h in headers})

# ── compute per-row metrics ───────────────────────────────────────────────────
USE_CASES = ['ai360_action_plan_emotional_bonding',
             'ai360_action_plan_task',
             'ai360_action_plan_product_storytelling']

UC_LABELS = {
    'ai360_action_plan_emotional_bonding':    'Emotional Bonding',
    'ai360_action_plan_task':                 'Task',
    'ai360_action_plan_product_storytelling': 'Product Storytelling',
}

# ── read total counts from original full dataset ──────────────────────────────
import json as _json
from collections import Counter as _Counter
_wb_orig = openpyxl.load_workbook(SRC_ORIG)
_ws_orig = _wb_orig['Sheet1']
_oh = [_ws_orig.cell(1,c).value for c in range(1, _ws_orig.max_column+1)]
_oc = {h: i+1 for i,h in enumerate(_oh)}
TOTAL_BY_UC  = _Counter()
EDITED_BY_UC = _Counter()

# per-SA counters: total messages and edited messages, overall and per use_case
_sa_total        = _Counter()   # seller_id → total msgs
_sa_edited       = _Counter()   # seller_id → edited msgs
_sa_total_uc     = {}           # (seller_id, uc) → total msgs
_sa_edited_uc    = {}           # (seller_id, uc) → edited msgs

for _r in range(2, _ws_orig.max_row+1):
    _uc  = _ws_orig.cell(_r, _oc['use_case']).value
    _sid = _ws_orig.cell(_r, _oc['seller_id']).value
    TOTAL_BY_UC[_uc] += 1
    _sa_total[_sid]  += 1
    _sa_total_uc[(_sid, _uc)] = _sa_total_uc.get((_sid, _uc), 0) + 1
    _cth = _ws_orig.cell(_r, _oc['copy_to_chat_history']).value
    if _cth:
        try:
            _d = _json.loads(_cth)
            _lv = _d.get(_d.get('latest_version','v1'), {})
            if _lv.get('is_change'):
                EDITED_BY_UC[_uc]       += 1
                _sa_edited[_sid]        += 1
                _sa_edited_uc[(_sid, _uc)] = _sa_edited_uc.get((_sid, _uc), 0) + 1
        except: pass
_wb_orig.close()
TOTAL_ALL  = sum(TOTAL_BY_UC.values())
EDITED_ALL = sum(EDITED_BY_UC.values())

# ── per-SA edit rates ─────────────────────────────────────────────────────────
import statistics as _stats

def _sa_rates(uc=None):
    """Return list of each SA's edit rate (%) for the given use_case (or overall)."""
    rates = []
    sids = set(_sa_total.keys())
    for sid in sids:
        if uc is None:
            total  = _sa_total[sid]
            edited = _sa_edited[sid]
        else:
            total  = _sa_total_uc.get((sid, uc), 0)
            edited = _sa_edited_uc.get((sid, uc), 0)
        if total > 0:
            rates.append(edited / total * 100)
    return rates

def _sa_distribution_kpis(uc=None, min_msgs=0):
    sids = set(_sa_total.keys())
    rates = []
    for sid in sids:
        if uc is None:
            total  = _sa_total[sid]
            edited = _sa_edited[sid]
        else:
            total  = _sa_total_uc.get((sid, uc), 0)
            edited = _sa_edited_uc.get((sid, uc), 0)
        if total >= max(min_msgs, 1):
            rates.append(edited / total * 100)
    n = len(rates)
    if n == 0:
        return {}
    mean_rate   = round(_stats.mean(rates), 1)
    median_rate = round(_stats.median(rates), 1)
    b0   = round(sum(1 for r in rates if r == 0)        / n * 100, 1)
    b1   = round(sum(1 for r in rates if 0 < r <= 20)   / n * 100, 1)
    b2   = round(sum(1 for r in rates if 20 < r <= 75)  / n * 100, 1)
    b3   = round(sum(1 for r in rates if r > 75)        / n * 100, 1)
    return {
        'sa_count':           n,
        'mean_sa_rate':       f"{mean_rate}%",
        'median_sa_rate':     f"{median_rate}%",
        'bucket_0':           f"{b0}%  ({sum(1 for r in rates if r == 0)} SAs)",
        'bucket_1_20':        f"{b1}%  ({sum(1 for r in rates if 0 < r <= 20)} SAs)",
        'bucket_20_75':       f"{b2}%  ({sum(1 for r in rates if 20 < r <= 75)} SAs)",
        'bucket_75plus':      f"{b3}%  ({sum(1 for r in rates if r > 75)} SAs)",
    }

# Median messages per SA = minimum threshold for filtered distribution
_all_msg_counts  = list(_sa_total.values())
MEDIAN_MSG_THRESHOLD = int(_stats.median(_all_msg_counts))

SA_DIST_OVERALL        = _sa_distribution_kpis()
SA_DIST_BY_UC          = {uc: _sa_distribution_kpis(uc) for uc in USE_CASES}
SA_DIST_FILTERED        = _sa_distribution_kpis(min_msgs=MEDIAN_MSG_THRESHOLD)
SA_DIST_FILTERED_BY_UC  = {uc: _sa_distribution_kpis(uc, min_msgs=MEDIAN_MSG_THRESHOLD) for uc in USE_CASES}

# SA message volume distribution (overall only — explains skew)
_counts_desc   = sorted(_all_msg_counts, reverse=True)
_n_sa          = len(_counts_desc)
_total_msgs    = sum(_counts_desc)
_low_vol_n     = sum(1 for x in _counts_desc if 1 <= x <= 5)
_low_vol_pct   = round(_low_vol_n / _n_sa * 100, 1)
_top5_msgs     = sum(_counts_desc[:5])
_top5_pct      = round(_top5_msgs / _total_msgs * 100, 1)
_top5_min      = _counts_desc[4]
_top5_max      = _counts_desc[0]
_mean_msgs     = round(_stats.mean(_counts_desc), 1)
_median_msgs   = round(_stats.median(_counts_desc), 1)
_ratio         = round(_mean_msgs / _median_msgs, 1)

SA_VOL = {
    'vol_low_volume':    f"{_low_vol_pct}%  ({_low_vol_n} of {_n_sa} SAs)",
    'vol_top5':          f"Top 5 SAs → {_top5_pct}% of all messages  ({_top5_min}–{_top5_max} msgs each)",
    'vol_mean':          f"{_mean_msgs}",
    'vol_median':        f"{_median_msgs}",
    'vol_ratio':         f"Mean is {_ratio}x the median",
}

def classify(rec):
    summary = rec['change_summary']
    ai_en   = rec['ai_message_translated']
    sent_en = rec['sent_message_translated']
    ai_orig = rec['ai_message_original']
    sent_orig = rec['sent_message_original']

    added_en   = added_segments(ai_en,   sent_en)
    removed_en = removed_segments(ai_en, sent_en)
    added_orig = added_segments(ai_orig, sent_orig)

    # length change bucket — Chinese character counts (threshold: 25%)
    _ai_wc_raw   = len(ai_orig)
    _sent_wc_raw = len(sent_orig)
    _pct = ((_sent_wc_raw - _ai_wc_raw) / _ai_wc_raw * 100) if _ai_wc_raw else 0
    if _pct <= -25:
        length = 'Shortened'
    elif _pct >= 25:
        length = 'Expanded'
    elif abs(_pct) >= 10:
        length = 'Moderate'
    else:
        length = 'Minor'

    # location
    loc_line = ""
    for line in summary.splitlines():
        if line.startswith("Location:"):
            loc_line = line.lower()
    in_opening = 'opening' in loc_line
    in_middle  = 'middle'  in loc_line
    in_closing = 'closing' in loc_line
    whole      = 'entire'  in loc_line or 'whole' in loc_line
    if whole:
        in_opening = in_middle = in_closing = True

    # emoji
    emoji_added   = has_emoji(added_en) or has_emoji(added_orig)
    emoji_removed = has_emoji(removed_en)

    # Chinese diff text (all changed characters, added + removed)
    removed_orig   = removed_segments(ai_orig, sent_orig)
    diff_orig_all  = added_orig + removed_orig   # full changed text in Chinese

    # Jieba-derived terms + domain knowledge for Chinese luxury retail
    CN_GREET   = ['亲爱', '您好', '你好', '小姐', '姐姐', '女士', '好久不见',
                  '先生', '太太', '早上好', '下午好', '晚上好', '嗨', '哈喽', '久违']
    CN_CTA     = ['有空', '看看', '联系', '随时', '过来', '试试', '试戴', '带来', '方便',
                  '来店', '到店', '欢迎', '预约', '安排', '拜访', '有机会', '欢迎光临']
    CN_PRODUCT = ['珠宝', '卡地亚', '系列', '腕表', '戒指', 'LOVE', 'Love',
                  '猎豹', '玫瑰', '钻石', 'Juste', 'Clou', '手镯', '项链',
                  '手表', '吊坠', '耳环', '胸针', '蓝气球', '三环', 'Trinity',
                  '钉子', 'Santos', 'Tank', 'Clash', '宝石', '项圈']
    CN_SCARCITY= ['入手', '购买', '现在', '库存', '现货', '稀缺', '有限', '紧张', '难得', '最后']
    CN_CARE    = ['清洗', '服务', '清洁', '保养', '免费', '护理', '每年',
                  '维修', '抛光', '售后', '养护', '定期']
    CN_SEASON  = ['下午', '下个月', '去年', '上次',
                  '春', '夏', '秋', '冬', '节日', '新年', '圣诞', '礼物', '礼品',
                  '情人节', '母亲节', '七夕', '节庆', '假期', '生日', '周年']

    greeting_mod    = any(w in diff_orig_all  for w in CN_GREET)
    cta_modified    = any(w in diff_orig_all  for w in CN_CTA)
    product_modified= any(w in diff_orig_all  for w in CN_PRODUCT)
    scarcity        = any(w in diff_orig_all  for w in CN_SCARCITY)
    aftersales      = any(w in diff_orig_all  for w in CN_CARE)
    seasonal        = any(w in diff_orig_all  for w in CN_SEASON)

    # Chinese character counts
    ai_chars   = len(ai_orig)
    sent_chars = len(sent_orig)

    return {
        'use_case':         rec['use_case'],
        'length':           length,
        'in_opening':       in_opening,
        'in_middle':        in_middle,
        'in_closing':       in_closing,
        'emoji_added':      emoji_added,
        'emoji_removed':    emoji_removed,
        'cta_modified':     cta_modified,
        'greeting_mod':     greeting_mod,
        'product_modified': product_modified,
        'scarcity':         scarcity,
        'aftersales':       aftersales,
        'seasonal':         seasonal,
        'ai_chars':         ai_chars,
        'sent_chars':       sent_chars,
    }

metrics = [classify(r) for r in records]

# ── aggregate ─────────────────────────────────────────────────────────────────
def agg(subset):
    n = len(subset)
    if n == 0:
        return {}

    def pct(flag): return round(sum(1 for m in subset if m[flag]) / n * 100, 1)
    def pct_count(*buckets): return round(sum(1 for m in subset if m['length'] in buckets) / n * 100, 1)
    def avg_chars(key, bucket=None):
        rows = [m for m in subset if bucket is None or m['length'] == bucket]
        vals = [m[key] for m in rows if m[key] is not None]
        return round(sum(vals) / len(vals)) if vals else '—'

    shortened = pct_count('Shortened')
    expanded  = pct_count('Expanded')

    return {
        'n':                        n,
        'pct_shortened':            f"{shortened}%",
        'pct_expanded':             f"{expanded}%",
        'pct_no_change':            f"{pct_count('Minor')}%",
        'avg_ai_chars':             avg_chars('ai_chars'),
        'avg_sent_chars':           avg_chars('sent_chars'),
        'avg_ai_chars_shortened':   avg_chars('ai_chars',  'Shortened'),
        'avg_sent_chars_shortened': avg_chars('sent_chars','Shortened'),
        'avg_ai_chars_expanded':    avg_chars('ai_chars',  'Expanded'),
        'avg_sent_chars_expanded':  avg_chars('sent_chars','Expanded'),
        'pct_opening':              f"{pct('in_opening')}%",
        'pct_middle':               f"{pct('in_middle')}%",
        'pct_closing':              f"{pct('in_closing')}%",
        'pct_emoji_added':          f"{pct('emoji_added')}%",
        'pct_emoji_removed':        f"{pct('emoji_removed')}%",
        'pct_cta':                  f"{pct('cta_modified')}%",
        'pct_greeting':             f"{pct('greeting_mod')}%",
        'pct_product_modified':     f"{pct('product_modified')}%",
        'pct_scarcity':             f"{pct('scarcity')}%",
        'pct_aftersales':           f"{pct('aftersales')}%",
        'pct_seasonal':             f"{pct('seasonal')}%",
    }

overall = agg(metrics)
by_uc   = {uc: agg([m for m in metrics if m['use_case'] == uc]) for uc in USE_CASES}

# Inject edit-rate fields (need the full-dataset totals, not just the edited subset)
def edit_rate_fields(uc=None):
    if uc is None:
        total  = TOTAL_ALL
        edited = EDITED_ALL
    else:
        total  = TOTAL_BY_UC.get(uc, 0)
        edited = EDITED_BY_UC.get(uc, 0)
    rate = round(edited / total * 100, 1) if total else 0
    return {'total_generated': total, 'total_edited': edited,
            'edit_rate': f"{rate}%  ({edited} / {total})"}

overall['total_generated'] = edit_rate_fields()['total_generated']
overall['total_edited']    = edit_rate_fields()['total_edited']
overall['edit_rate']       = edit_rate_fields()['edit_rate']
for uc in USE_CASES:
    ef = edit_rate_fields(uc)
    by_uc[uc]['total_generated'] = ef['total_generated']
    by_uc[uc]['total_edited']    = ef['total_edited']
    by_uc[uc]['edit_rate']       = ef['edit_rate']

# Inject SA distribution fields (unfiltered)
overall.update(SA_DIST_OVERALL)
for uc in USE_CASES:
    by_uc[uc].update(SA_DIST_BY_UC[uc])

# Inject SA volume insights (overall only — by_uc will show '—' by default)
overall.update(SA_VOL)


# ── KPI rows definition ───────────────────────────────────────────────────────
SECTIONS = [
    ("── EDIT RATE ──", None),
    ("Total AI-generated messages",          'total_generated'),
    ("Total SA-edited messages",             'total_edited'),
    ("Edit rate  (edited / total generated)", 'edit_rate'),

    ("── LENGTH CHANGE  (Chinese character count) ──", None),
    ("% Shortened  (> 25% fewer characters)",          'pct_shortened'),
    ("Avg characters — AI message (shortened rows)",   'avg_ai_chars_shortened'),
    ("Avg characters — Sent message (shortened rows)", 'avg_sent_chars_shortened'),
    ("% Expanded   (> 25% more characters)",           'pct_expanded'),
    ("Avg characters — AI message (expanded rows)",    'avg_ai_chars_expanded'),
    ("Avg characters — Sent message (expanded rows)",  'avg_sent_chars_expanded'),
    ("% No significant change  (< 10%)",               'pct_no_change'),
    ("Avg characters — AI message (all)",              'avg_ai_chars'),
    ("Avg characters — Sent message (all)",            'avg_sent_chars'),

    ("── LOCATION OF CHANGE ──", None),
    ("% with edits in opening section",  'pct_opening'),
    ("% with edits in middle section",   'pct_middle'),
    ("% with edits in closing section",  'pct_closing'),

    ("── CONTENT TYPE ──", None),
    ("% modified greeting / salutation", 'pct_greeting'),
    ("% modified call-to-action",        'pct_cta'),
    ("% modified product references",    'pct_product_modified'),
    ("% added emoji",                    'pct_emoji_added'),
    ("% removed emoji from AI",          'pct_emoji_removed'),

    ("── SA DISTRIBUTION ──", None),
    ("Total SAs with AI-generated messages",         'sa_count'),
    ("% of SAs who never edited  (0%)",              'bucket_0'),
    ("% of SAs with low edit rate  (1–20%)",         'bucket_1_20'),
    ("% of SAs with moderate edit rate  (20–75%)",  'bucket_20_75'),
    ("% of SAs with high edit rate  (> 75%)",       'bucket_75plus'),
    ("Mean edit rate per SA",                        'mean_sa_rate'),
    ("Median edit rate per SA",                      'median_sa_rate'),

]

# ── write KPI-only workbook ───────────────────────────────────────────────────
wb_out = xlsxwriter.Workbook(OUT)
ws_kpi  = wb_out.add_worksheet("KPI Summary")
wb_out.add_worksheet("Notes")

# ── KPI sheet ─────────────────────────────────────────────────────────────────
cols = ['Overall'] + [UC_LABELS[uc] for uc in USE_CASES]
agg_data = [overall] + [by_uc[uc] for uc in USE_CASES]

# formats
title_fmt  = wb_out.add_format({'bold':True,'font_size':14,'font_color':'#1F3864','bottom':2})
section_fmt= wb_out.add_format({'bold':True,'font_size':10,'bg_color':'#1F3864',
                                 'font_color':'#FFFFFF','text_wrap':True})
metric_fmt = wb_out.add_format({'font_size':10,'text_wrap':True,'bg_color':'#F4F6FB'})
metric_fmt2= wb_out.add_format({'font_size':10,'text_wrap':True,'bg_color':'#FFFFFF'})
num_fmt    = wb_out.add_format({'font_size':11,'bold':True,'align':'center',
                                 'bg_color':'#EEF2FF','border':1,'border_color':'#C0C8E0'})
num_fmt2   = wb_out.add_format({'font_size':11,'bold':True,'align':'center',
                                 'bg_color':'#FFFFFF','border':1,'border_color':'#C0C8E0'})
hdr_col_fmt= wb_out.add_format({'bold':True,'font_size':11,'bg_color':'#1F3864',
                                 'font_color':'#FFFFFF','align':'center','border':1,
                                 'border_color':'#C0C8E0','text_wrap':True})
kpi_title_fmt = wb_out.add_format({'bold':True,'font_size':10,'bg_color':'#F4F6FB',
                                    'text_wrap':True,'valign':'vcenter'})
kpi_title_fmt2= wb_out.add_format({'bold':True,'font_size':10,'bg_color':'#FFFFFF',
                                    'text_wrap':True,'valign':'vcenter'})

ws_kpi.set_column(0, 0, 40)   # metric label
ws_kpi.set_column(1, len(cols), 20)  # value columns

# Title
ws_kpi.merge_range(0, 0, 0, len(cols), "SA Edit Pattern Analysis — KPI Summary", title_fmt)
ws_kpi.set_row(0, 28)

# Column headers
ws_kpi.write(1, 0, "KPI Metric", hdr_col_fmt)
for ci, c in enumerate(cols, 1):
    ws_kpi.write(1, ci, c, hdr_col_fmt)
ws_kpi.set_row(1, 30)

# Data rows
kpi_row = 2
metric_parity = 0
for label, key in SECTIONS:
    if key is None:
        # section header row
        ws_kpi.merge_range(kpi_row, 0, kpi_row, len(cols), label, section_fmt)
        ws_kpi.set_row(kpi_row, 18)
        kpi_row += 1
        metric_parity = 0
        continue

    even = metric_parity % 2 == 0
    lf = kpi_title_fmt if even else kpi_title_fmt2
    vf = num_fmt if even else num_fmt2

    ws_kpi.write(kpi_row, 0, label, lf)
    for ci, ag in enumerate(agg_data, 1):
        val = ag.get(key, '—')
        ws_kpi.write(kpi_row, ci, val if val is not None else '—', vf)
    ws_kpi.set_row(kpi_row, 20)
    kpi_row += 1
    metric_parity += 1

# freeze top rows
ws_kpi.freeze_panes(2, 1)

wb_out.close()
print(f"Done — {OUT}")
