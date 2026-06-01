import difflib
import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Font, PatternFill, Alignment, Color

FILE = "SA_edits_translated.xlsx"

# ── fonts for rich text ──────────────────────────────────────────────────────
F_NORMAL = InlineFont(sz=11)
F_RED    = InlineFont(sz=11, color="CC0000")          # removed from AI msg
F_BOLD   = InlineFont(sz=11, b=True)                  # added in sent msg

# ── diff helpers ─────────────────────────────────────────────────────────────
def char_diff(a, b):
    """
    Returns (segs_a, segs_b) where each is a list of (tag, text).
    tag in {'equal', 'delete', 'insert'}
    segs_a → what to show in the ai_message cell  (red = deleted)
    segs_b → what to show in the sent_message cell (bold = added)
    """
    sm = difflib.SequenceMatcher(None, a, b, autojunk=False)
    segs_a, segs_b = [], []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == 'equal':
            segs_a.append(('equal',  a[i1:i2]))
            segs_b.append(('equal',  b[j1:j2]))
        elif tag == 'delete':
            segs_a.append(('delete', a[i1:i2]))
        elif tag == 'insert':
            segs_b.append(('insert', b[j1:j2]))
        elif tag == 'replace':
            segs_a.append(('delete', a[i1:i2]))
            segs_b.append(('insert', b[j1:j2]))
    return segs_a, segs_b

def to_rich_ai(segs):
    """ai_message cell: normal text + red for deleted parts"""
    blocks = []
    for tag, text in segs:
        blocks.append(TextBlock(F_RED if tag == 'delete' else F_NORMAL, text))
    return CellRichText(blocks) if blocks else ""

def to_rich_sent(segs):
    """sent_message cell: normal text + bold for inserted parts"""
    blocks = []
    for tag, text in segs:
        blocks.append(TextBlock(F_BOLD if tag == 'insert' else F_NORMAL, text))
    return CellRichText(blocks) if blocks else ""

# ── summary generator ────────────────────────────────────────────────────────
def summarise(ai_en, sent_en):
    """
    Produce a plain-English summary of what changed, useful for prompt tuning.
    Works on the English-translated versions for readability.
    """
    ai_words   = ai_en.split()
    sent_words = sent_en.split()
    n_ai, n_sent = len(ai_words), len(sent_words)

    removed_chunks, added_chunks = [], []
    sm = difflib.SequenceMatcher(None, ai_words, sent_words, autojunk=False)
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag in ('delete', 'replace'):
            removed_chunks.append(' '.join(ai_words[i1:i2]))
        if tag in ('insert', 'replace'):
            added_chunks.append(' '.join(sent_words[j1:j2]))

    # ── length change ──
    diff_words = n_sent - n_ai
    pct = round(diff_words / n_ai * 100) if n_ai else 0
    if pct <= -30:
        length_note = f"Significantly shortened ({abs(pct)}% fewer words)"
    elif pct < -10:
        length_note = f"Shortened ({abs(pct)}% fewer words)"
    elif pct >= 30:
        length_note = f"Significantly expanded ({pct}% more words)"
    elif pct > 10:
        length_note = f"Expanded ({pct}% more words)"
    else:
        length_note = f"Similar length ({'+' if diff_words>=0 else ''}{diff_words} words)"

    # ── where did the change happen? ──
    n_chars = max(len(ai_en), len(sent_en))
    sm_char = difflib.SequenceMatcher(None, ai_en, sent_en, autojunk=False)
    changed_positions = []
    for tag, i1, i2, j1, j2 in sm_char.get_opcodes():
        if tag != 'equal':
            changed_positions.append((i1 + i2) / 2)   # midpoint of change in ai_en

    zones = {'opening': 0, 'middle': 0, 'closing': 0}
    for pos in changed_positions:
        rel = pos / n_chars if n_chars else 0.5
        if rel < 0.30:
            zones['opening'] += 1
        elif rel > 0.70:
            zones['closing'] += 1
        else:
            zones['middle'] += 1

    dominant = max(zones, key=zones.get) if any(zones.values()) else 'whole message'
    if all(v == 0 for v in zones.values()):
        zone_note = "whole message rewritten"
    elif sum(1 for v in zones.values() if v > 0) == 3:
        zone_note = "changes throughout entire message"
    elif sum(1 for v in zones.values() if v > 0) == 2:
        parts = [k for k, v in zones.items() if v > 0]
        zone_note = f"changes in {' & '.join(parts)}"
    else:
        zone_note = f"main change in {dominant}"

    # ── content clues ──
    content_notes = []
    all_removed = ' '.join(removed_chunks).lower()
    all_added   = ' '.join(added_chunks).lower()

    if any(w in all_removed for w in ['hello', 'dear', 'hi ', 'miss', 'mr', 'ms', 'madam']):
        content_notes.append("modified greeting/salutation")
    if any(w in all_added for w in ['hello', 'dear', 'hi ', 'miss', 'mr', 'ms', 'madam']):
        content_notes.append("personalised salutation")
    if any(w in all_removed for w in ['please', 'come', 'visit', 'welcome', 'contact', 'reach', 'appointment', 'arrange']):
        content_notes.append("modified call-to-action")
    if any(w in all_added for w in ['please', 'come', 'visit', 'welcome', 'contact', 'reach', 'appointment', 'arrange']):
        content_notes.append("added/changed call-to-action")
    if any(w in all_added for w in ['love', 'juste un clou', 'trinity', 'panthère', 'santos', 'tank', 'ballon', 'clash', 'bracelet', 'ring', 'necklace', 'watch']):
        content_notes.append("added product name/series reference")
    if any(w in all_removed for w in ['love', 'juste un clou', 'trinity', 'panthère', 'santos', 'tank', 'ballon', 'clash', 'bracelet', 'ring', 'necklace', 'watch']):
        content_notes.append("removed product reference")
    if any(w in all_added for w in ['stock', 'inventory', 'limited', 'available', 'rare', 'scarce', 'last']):
        content_notes.append("added scarcity/urgency")
    if any(w in all_removed for w in ['stock', 'inventory', 'limited', 'available', 'rare', 'scarce']):
        content_notes.append("removed scarcity language")
    if any(w in all_added for w in ['clean', 'maintain', 'service', 'repair', 'polish', 'care']):
        content_notes.append("added after-sales/care mention")
    if any(w in all_added for w in ['spring', 'summer', 'winter', 'autumn', 'holiday', 'festival', 'new year', 'christmas', 'gift']):
        content_notes.append("added seasonal/occasion reference")
    if n_sent > 0 and n_ai > 0 and len(sent_chunks := [c for c in added_chunks if len(c) > 60]) > 0:
        content_notes.append("added substantial new content")

    # ── removed & added excerpts ──
    removed_preview = "; ".join(f'"{c}"' for c in removed_chunks if c.strip())[:300]
    added_preview   = "; ".join(f'"{c}"' for c in added_chunks if c.strip())[:300]

    parts = [length_note, f"Location: {zone_note}"]
    if content_notes:
        parts.append("Type: " + ", ".join(dict.fromkeys(content_notes)))   # dedup order-preserving
    if removed_preview:
        parts.append(f"Removed: {removed_preview}")
    if added_preview:
        parts.append(f"Added: {added_preview}")

    return "\n".join(parts)

# ── main ─────────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(FILE, rich_text=True)
ws = wb.active

# Read current headers
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
print("Current columns:", headers)

# Column indices (1-based)
COL = {h: i+1 for i, h in enumerate(headers)}
AI_ORIG   = COL['ai_message_original']
AI_EN     = COL['ai_message_translated']
SENT_ORIG = COL['sent_message_original']
SENT_EN   = COL['sent_message_translated']

# Add new header: change_summary (after sent_message_translated)
NEW_COL = ws.max_column + 1
header_fill = PatternFill("solid", fgColor="1F3864")
header_font = Font(bold=True, color="FFFFFF", size=11)
cell = ws.cell(1, NEW_COL, "change_summary")
cell.fill = header_fill
cell.font = header_font
cell.alignment = Alignment(wrap_text=True, vertical='center')

fill_a = PatternFill("solid", fgColor="EEF2FF")
fill_b = PatternFill("solid", fgColor="FFFFFF")
wrap   = Alignment(wrap_text=True, vertical='top')

total = ws.max_row - 1
for row in range(2, ws.max_row + 1):
    i = row - 2
    if (i + 1) % 20 == 0 or i == 0:
        print(f"  [{i+1}/{total}] processing row {row}...")

    ai_orig_val   = ws.cell(row, AI_ORIG).value   or ""
    ai_en_val     = ws.cell(row, AI_EN).value     or ""
    sent_orig_val = ws.cell(row, SENT_ORIG).value or ""
    sent_en_val   = ws.cell(row, SENT_EN).value   or ""

    # Strip any existing rich text to plain string
    def plain(v):
        if isinstance(v, CellRichText):
            return "".join(str(t) for t in v)
        return str(v) if v else ""

    ai_orig_s   = plain(ai_orig_val)
    ai_en_s     = plain(ai_en_val)
    sent_orig_s = plain(sent_orig_val)
    sent_en_s   = plain(sent_en_val)

    # ── compute diffs ──
    segs_ai_orig,   segs_sent_orig = char_diff(ai_orig_s, sent_orig_s)
    segs_ai_en,     segs_sent_en   = char_diff(ai_en_s,   sent_en_s)

    # ── build rich text ──
    rich_ai_orig   = to_rich_ai(segs_ai_orig)
    rich_ai_en     = to_rich_ai(segs_ai_en)
    rich_sent_orig = to_rich_sent(segs_sent_orig)
    rich_sent_en   = to_rich_sent(segs_sent_en)

    fill = fill_a if i % 2 == 0 else fill_b

    def set_cell(r, c, val):
        cell = ws.cell(r, c)
        cell.value = val
        cell.fill  = fill
        cell.alignment = wrap

    set_cell(row, AI_ORIG,   rich_ai_orig)
    set_cell(row, AI_EN,     rich_ai_en)
    set_cell(row, SENT_ORIG, rich_sent_orig)
    set_cell(row, SENT_EN,   rich_sent_en)

    # ── change summary ──
    summary = summarise(ai_en_s, sent_en_s)
    sum_cell = ws.cell(row, NEW_COL, summary)
    sum_cell.fill      = fill
    sum_cell.alignment = wrap

# Widen change_summary column
ws.column_dimensions[ws.cell(1, NEW_COL).column_letter].width = 60

wb.save(FILE)
print(f"\nDone — saved to {FILE}")
